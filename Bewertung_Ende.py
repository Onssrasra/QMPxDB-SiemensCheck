import pandas as pd
import re
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ================= Pfade: Ein-/Ausgabe im gleichen Ordner =================
try:
    SCRIPT_DIR = Path(__file__).parent.resolve()
except NameError:
    # Falls im Notebook/REPL: aktuelles Arbeitsverzeichnis
    SCRIPT_DIR = Path.cwd().resolve()

input_path = SCRIPT_DIR / "Herstellerdaten für Clickbot-20-05-2025.xlsx"
output_path = SCRIPT_DIR / "Clickbot_Bewertung_Ergebnis.xlsx"

# ================= 1) Einlesen =================
# Überspringe 2 Kopfzeilen wie im Original
df = pd.read_excel(input_path, skiprows=2)

# ================= 2) Gültige Werte für Fert./Prüfhinweis =================
pos1_values = {"OHNE", "1", "2", "3"}
pos2_values = {"N", "3.2", "3.1", "2.2", "2.1"}
pos3_values = {"N", "CL1", "CL2", "CL3"}
pos4_values = {"N", "J"}
pos5_values = {"N", "A1", "A2", "A3", "A5", "A+"}

# ================= 3) Prüfung: Fert./Prüfhinweis vollständig und gültig? =================
def check_vollstaendig(hinweis):
    # NaN → Fehler
    if pd.isna(hinweis):
        return "1"
    # Exakt 5 Teile erwartet
    teile = [teil.strip() for teil in str(hinweis).split("/")]
    if len(teile) != 5:
        return "1"
    # Alle 5 Positionen müssen gültig sein
    return (
        "0"
        if (teile[0] in pos1_values and teile[1] in pos2_values and
            teile[2] in pos3_values and teile[3] in pos4_values and
            teile[4] in pos5_values)
        else "1"
    )

df["Fehler_Vollständigkeit_Fert./Prüfhinweis"] = df["Fert./Prüfhinweis"].apply(check_vollstaendig)

# ================= 4) Prüfung: Pflichtfelder B–J, N, R–W vollständig? =================
def check_bis_n_vollstaendig(row):
    # Spaltenindizes: A=0, B=1,..., J=9; N=13; R=17,..., W=22 → range(17, 23)
    relevant_indices = list(range(1, 10)) + [13] + list(range(17, 23))
    for i in relevant_indices:
        col = df.columns[i]
        if pd.isna(row[col]) or str(row[col]).strip() == "":
            return 1
    return 0

df["Fehler_Vollständigkeit_B-J+N+R-W"] = df.apply(check_bis_n_vollstaendig, axis=1)

# ================= 5) Maßprüfung =================
def check_masspruefung(row):
    # Robuste Konvertierung + Fallback
    try:
        l = float(row["Länge"]) if not pd.isna(row["Länge"]) else 0
        b = float(row["Breite"]) if not pd.isna(row["Breite"]) else 0
        h = float(row["Höhe"]) if not pd.isna(row["Höhe"]) else 0
        txt = str(row["Materialkurztext"]) if not pd.isna(row["Materialkurztext"]) else ""
    except Exception:
        return 1

    # Erkenne Textmaße wie 12×34 / 12x34 / 12 X 34 / 12*34 / 12/34
    pattern = r"\d{1,4}[\s×xX*/]{1,3}\d{1,4}"
    has_text_measurement = bool(re.search(pattern, txt))

    # Wenn L, B, H alle 0 sind, muss es zumindest ein Textmaß geben
    if l == 0 and b == 0 and h == 0:
        return 0 if has_text_measurement else 1
    return 0

df["Fehler_Maßprüfung"] = df.apply(check_masspruefung, axis=1)

# ================= 6) Gesamtfehler =================
# Max über die drei Fehlerarten → 1, sobald irgendwo ein Fehler ist
df["Fehler"] = df[[
    "Fehler_Vollständigkeit_Fert./Prüfhinweis",
    "Fehler_Vollständigkeit_B-J+N+R-W",
    "Fehler_Maßprüfung"
]].astype(int).max(axis=1)

# ================= 7) Aufteilen in zwei Blätter =================
df_ok   = df[df["Fehler"] == 0].copy()
df_bad  = df[df["Fehler"] != 0].copy()

# ================= 8) Export (zwei Sheets): Ohne_Fehler / Mit_Fehlern =================
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    df_ok.to_excel(writer,  index=False, sheet_name="Ohne_Fehler")
    df_bad.to_excel(writer, index=False, sheet_name="Mit_Fehlern")

# ================= 9) Nachträglich: Fehlerzellen (Wert "1") rot färben =================
wb = load_workbook(output_path)
fill_red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

error_cols = [
    "Fehler_Vollständigkeit_Fert./Prüfhinweis",
    "Fehler_Vollständigkeit_B-J+N+R-W",
    "Fehler_Maßprüfung",
    "Fehler"
]

def color_errors(ws):
    # Kopfzeile lesen → Spaltenindex robust ermitteln
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = [header.index(col) + 1 for col in error_cols if col in header]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for idx in col_idx:
            cell = row[idx - 1]
            if str(cell.value).strip() == "1":
                cell.fill = fill_red

if "Ohne_Fehler" in wb.sheetnames:
    color_errors(wb["Ohne_Fehler"])
if "Mit_Fehlern" in wb.sheetnames:
    color_errors(wb["Mit_Fehlern"])

# ================= 10) Drittes Blatt: Zusammenfassung/Qualitätsprüfung =================
# — Statistik robust mit numerischer Umwandlung (0/1) —
def col_sum_and_rate(frame, colname):
    total = len(frame)
    s = pd.to_numeric(frame[colname], errors="coerce").fillna(0).astype(int)
    cnt = int(s.sum())
    rate = round((cnt / total) * 100, 2) if total > 0 else 0.0
    return cnt, rate

gesamt_cnt, gesamt_rate = col_sum_and_rate(df, "Fehler")
pf_cnt, pf_rate         = col_sum_and_rate(df, "Fehler_Vollständigkeit_B-J+N+R-W")
pr_cnt, pr_rate         = col_sum_and_rate(df, "Fehler_Vollständigkeit_Fert./Prüfhinweis")
mass_cnt, mass_rate     = col_sum_and_rate(df, "Fehler_Maßprüfung")

# Neues Blatt anlegen und formatieren
ws_sum = wb.create_sheet(title="Qualitätsprüfung")
header_font = Font(bold=True)
header_fill = PatternFill("solid", fgColor="FFFF99")
center = Alignment(horizontal="center")

# Kopfzeile + Daten (analog deiner Vorlage)
header1 = ["", "Fehleranzahl", "Fehlerquote"]
data1 = [
    [f"Gesamt({len(df)})", gesamt_cnt, f"{gesamt_rate}%"],
    ["Vollständigkeit_Pflichtfeld", pf_cnt, f"{pf_rate}%"],
    ["Vollständigkeit_Fert./Prüfhinweis", pr_cnt, f"{pr_rate}%"],
    ["Gültigkeit_Maß", mass_cnt, f"{mass_rate}%"]
]

# Schreiben
ws_sum.append(header1)
for row in data1:
    ws_sum.append(row)

# Leere Zeile nach dem Block (optional, für Optik)
ws_sum.append([""] * 3)

# Zellen formatieren
max_row = ws_sum.max_row
for row in ws_sum.iter_rows(min_row=1, max_row=max_row, max_col=3):
    for cell in row:
        cell.alignment = center
        if cell.row == 1:  # Kopfzeile
            cell.font = header_font
            cell.fill = header_fill

# Optional: Spaltenbreite etwas anpassen
ws_sum.column_dimensions["A"].width = 35
ws_sum.column_dimensions["B"].width = 14
ws_sum.column_dimensions["C"].width = 14

# ================= 11) Speichern =================
wb.save(output_path)
print(f"✅ Datei mit 3 Sheets erstellt & Fehler farbig markiert: {output_path}")
